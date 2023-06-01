import tkinter
import tkinter.ttk
import tkinter.font
import tkinter.filedialog
import tkinter.messagebox

# check if openpyxl is installed
excel = False
try:
  import openpyxl
  excel = True
except ModuleNotFoundError:
  print("Openpyxl package is not found.")
  print("Loading Excel files is not supported.")


class DGA(object):
  """ performs dissolved gases analyses. """

  def __init__(self):
    self.root = tkinter.Tk() # create main window
    self.root.geometry("") # left blank for automatic scaling
    self.root.title("DGA using IEC 60599") # window title
    self.font_style = tkinter.font.Font(size=10) # style the fonts

    # create the labels
    self.heading = tkinter.Label(self.root,
                                 text="Diagnose Faults in Transformers",
                                 pady=15,
                                 font=self.font_style)
    self.h2 = tkinter.Label(self.root,
                            text="H2 (hydrogen)",
                            font=self.font_style)
    self.c2h2 = tkinter.Label(self.root,
                              text="C2H2 (acetylene)",
                              font=self.font_style)
    self.c2h4 = tkinter.Label(self.root,
                              text="C2H4 (ethylene)",
                              font=self.font_style)
    self.c2h6 = tkinter.Label(self.root,
                              text="C2H6 (ethane)",
                              font=self.font_style)
    self.ch4 = tkinter.Label(self.root,
                             text="CH4 (methane)",
                             font=self.font_style)

    # place the labels inside the window
    self.heading.grid(row=0, column=3)
    self.h2.grid(row=1, column=2)
    self.c2h2.grid(row=2, column=2)
    self.c2h4.grid(row=3, column=2)
    self.c2h6.grid(row=4, column=2)
    self.ch4.grid(row=5, column=2)

    # tkinter variables to store entered data
    self.h2_ppm = tkinter.StringVar()
    self.c2h2_ppm = tkinter.StringVar()
    self.c2h4_ppm = tkinter.StringVar()
    self.c2h6_ppm = tkinter.StringVar()
    self.ch4_ppm = tkinter.StringVar()

    # entries to enter the numbers
    self.h2_entry = tkinter.Entry(self.root,
                                  textvariable=self.h2_ppm,
                                  font=self.font_style)
    self.c2h2_entry = tkinter.Entry(self.root,
                                    textvariable=self.c2h2_ppm,
                                    font=self.font_style)
    self.c2h4_entry = tkinter.Entry(self.root,
                                    textvariable=self.c2h4_ppm,
                                    font=self.font_style)
    self.c2h6_entry = tkinter.Entry(self.root,
                                    textvariable=self.c2h6_ppm,
                                    font=self.font_style)
    self.ch4_entry = tkinter.Entry(self.root,
                                   textvariable=self.ch4_ppm,
                                   font=self.font_style)

    # place the entries inside the window
    self.h2_entry.grid(row=1, column=3)
    self.c2h2_entry.grid(row=2, column=3)
    self.c2h4_entry.grid(row=3, column=3)
    self.c2h6_entry.grid(row=4, column=3)
    self.ch4_entry.grid(row=5, column=3)

    # diagnose faults from an excel data
    if excel:
      self.load_button = tkinter.Button(self.root, width="10", text="Load File", command=self.open_file, font=self.font_style)
      self.load_button.grid(row=6, column=2)

    # diagnose faults from user input
    self.calculate_button = tkinter.Button(self.root, width=15, text="Submit Values", command=self.diagnose_user_input, font=self.font_style)
    self.calculate_button.grid(row=6, column=3)

    # label
    self.iec_method = tkinter.StringVar()
    tkinter.ttk.Label(self.root, text="Select the DGA method:",
                      font = self.font_style).grid(column=2, row=7, padx=5, pady=10)

    self.selections = tkinter.ttk.Combobox(self.root, width=20,
                                           textvariable=self.iec_method)
    self.selections["values"] = ["IEC Ratio", "IEC Ratio (Extended)", "Duval Triangle"]

    self.selections.grid(column=3, row=7)
    self.selections.current(0)

    self.selected_iec_method = "IEC Ratio"
    self.selections.bind("<<ComboboxSelected>>", self.iec_selector)

    self.root.mainloop()


  def iec_selector(self, event):
    self.selected_iec_method = self.selections.get()


  def diagnose_user_input(self):
    """ performs DGA-IEC on user input. """

    num_gases = []
    str_gases = [self.h2_ppm, self.c2h2_ppm, self.c2h4_ppm,
                 self.c2h6_ppm, self.ch4_ppm]

    # convert user input from string-type to float-type
    for gas in str_gases:
      num_gases.append(float(gas.get()))

    # call IEC test on user data
    if self.selected_iec_method == "IEC Ratio":
      diagnosis = self.iec(num_gases)
    elif self.selected_iec_method == "IEC Ratio (Extended)":
      diagnosis = self.iec_extended(num_gases)
    elif self.selected_iec_method == "Duval Triangle":
      diagnosis = self.duval(num_gases)

    # display IEC test results in the messagebox
    tkinter.messagebox.showinfo("Diagnosis", str(diagnosis))


  def open_file(self):
    """ loads an excel-like file and performs DGA-IEC on it. """

    data_file = tkinter.filedialog.askopenfile(title="Select a file", mode='r',
          filetypes =[("Excel Files", "*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv")])
    wb = openpyxl.load_workbook(filename=data_file.name)
    ws = wb.active

    self.diagnose_file_input(ws)


  def diagnose_file_input(self, ws):
    """ performs IEC on openpyxl object. """

    rows_iter = ws.iter_rows(min_col=1, min_row=1, max_col=10, max_row=ws.max_row)

    data = [[cell.value for cell in list(row)] for row in rows_iter]

    gas_idx = []
    iec_gases = ["h2", "c2h2", "c2h4", "c2h6", "ch4"]
    for gas in iec_gases:
      gas_idx.append(data[0].index(gas))

    ppms = []
    for row in data:
      tmp = []
      for idx in gas_idx:
        tmp.append(row[idx])
      ppms.append(tmp)

    faults = []
    if self.selected_iec_method == "IEC Ratio":
      for ppm in ppms[1:]:
        faults.append(self.iec(ppm))
    elif self.selected_iec_method == "IEC Ratio (Extended)":
      for ppm in ppms[1:]:
        faults.append(self.iec_extended(ppm))
    elif self.selected_iec_method == "Duval Triangle":
      for ppm in ppms[1:]:
        faults.append(self.duval(ppm))

    with open("faults.txt", "w") as out:
      # save to txt
      res = ['.'.join([str(x),y])
             for x,y in list(zip(range(1,len(faults)+1), faults))]
      res = '\n'.join(res)
      out.write(res)

      # display to user
      msg_win = tkinter.Tk()
      msg_win.geometry("600x150")
      msg_win.title("Diagnosed Faults")
      msg = tkinter.Scrollbar(msg_win)
      msg.pack(side=tkinter.RIGHT,fill=tkinter.Y)
      msg_list = tkinter.Listbox(msg_win,
                                 yscrollcommand=msg.set,
                                 background="lightgray",
                                 font=self.font_style)
      for idx, line in enumerate(faults):
        msg_list.insert(tkinter.END, str(idx) + ". " + line)
      msg_list.pack(side=tkinter.LEFT, fill="both", expand=True)
      msg.config(command=msg_list.yview)


  def duval(self, gases):
    diagnosis = None

    _, c2h2, c2h4, _, ch4 = gases

    total = ch4 + c2h4 + c2h2

    R1 = 100. * ch4 / total
    R2 = 100. * c2h4 / total
    R3 = 100. * c2h2 / total

    assert round(R1+R2+R3) == 100.

    if R1 >= 98 and R2 <= 2 and R3 <= 2:
      diagnosis = "Partial Discharge."
    elif R1 >= 46 and R1 <= 80 and R2 >= 20 and R2 <= 50 and R3 <= 4:
      diagnosis = "Thermal Fault: 300C < T <= 700C."
    elif R1 >= 76 and R1 <= 98 and R2 <= 20 and R3 <= 4:
      diagnosis = "Thermal Fault: T <= 300C."
    elif R2 <= 23 and R3 >= 0:
      diagnosis = "Discharge of Low Energy."
    elif R1 <= 50 and R2 >= 50 and R3 <= 15:
      diagnosis = "Thermal Fault: T > 700C."
    elif R2 <= 77 and R3 >= 13 and R3 <= 79:
      diagnosis = "Discharge of High Energy."
    elif R2 <= 85 and R3 <= 29 and R3 >= 4:
      diagnosis = "Discharge of High Energy."
    else:
      diagnosis = "Check entered values. Otherwise you can try using other DGA methods."

    return diagnosis


  def iec(self, gases):
    """ performs DGA using IEC 60599 ratio method. """

    diagnosis = None

    h2, c2h2, c2h4, c2h6, ch4 = gases

    R1 = c2h2 / c2h4
    R2 = ch4 / h2
    R3 = c2h4 / c2h6

    if R1 < 0.1 and R2 < 0.1 and R3 < 0.1:
      diagnosis = "No Faults."
    elif R1 < 0.1 and R2 > 1 and R3 >= 1 and R3 <= 4:
      diagnosis = "Thermal Fault: 300C < T <= 700C."
    elif R1 < 0.2 and R2 > 1 and R3 > 4:
      diagnosis = "Thermal Fault: T > 700C."
    elif R2 < 0.1 and R3 < 0.2:
      diagnosis = "Partial Discharge."
    elif R1 > 1 and R2 >= 0.1 and R2 <= 0.5 and R3 > 1:
      diagnosis = "Discharge of Low Energy."
    elif R1 >= 0.6 and R1 <= 2.5 and R2 >= 0.1 and R2 <= 1 and R3 > 2:
      diagnosis = "Discharge of High Energy."
    elif R2 > 1 and R3 < 1:
      diagnosis = "Thermal Fault: T <= 300C."
    else:
      diagnosis = "Check entered values. Otherwise you can try using other DGA methods."

    return diagnosis


  def iec_extended(self, gases):
    """
        performs DGA using extended IEC 60599 ratio method as descrbed in

        Shrivastava, K. and Choubey, A., 2012.
        A novel association rule mining with IEC ratio based
        dissolved gas analysis for fault diagnosis of power transformers.
        International journal of advanced computer research, 2(2), p.34.
    """

    diagnosis = None
    iec_code = [0,0,0]

    h2, c2h2, c2h4, c2h6, ch4 = gases

    R1 = c2h2 / c2h4
    R2 = ch4 / h2
    R3 = c2h4 / c2h6

    # <0.1: 0 1 0
    if R2 < 0.1:
      iec_code[1] = 1
    # 0.1-1: 1 0 0
    if (R1>=0.1 and R1<=3):
      iec_code[0] = 1
    # 1-3: 1 2 1
    if R2 >= 1:
      iec_code[1] = 2
    if (R3>=1 and R3<=3):
      iec_code[2] = 1
    # >3: 2 2 2
    if R1 > 3:
      iec_code[0] = 2
    if R3 > 3:
      iec_code[2] = 2

    iec_fault_codes = {
      "000":"No Fault.",
      "010":"Partial Discharge of Low Energy.",
      "110":"Partial Discharge of High Energy.",
      "101":"Discharge of Low Energy.",
      "202":"Discharge of Low Energy.",
      "102":"Discharge of High Energy.",
      "001":"Thermal Fault: T <= 150C.",
      "020":"Thermal Fault: 150C < T <= 300C.",
      "021":"Thermal Fault: 300C < T <= 700C.",
      "022":"Thermal Fault: T > 700C."
    }

    try:
      diagnosis = iec_fault_codes[''.join([str(x) for x in iec_code])]
    except:
      diagnosis = "Null"

    return diagnosis


if __name__ == '__main__':
  DGA()
